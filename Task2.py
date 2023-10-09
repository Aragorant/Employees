import csv
import openpyxl
from datetime import datetime

def calculate_age(birthdate):
    today = datetime.today()
    age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
    return age

try:
    with open('employees.csv', mode='r', newline='', encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file)

        wb = openpyxl.Workbook()

        if 'Sheet' in wb.sheetnames:
            sheet = wb['Sheet']
            wb.remove(sheet)

        fieldnames = ["Прізвище", "Ім’я", "По батькові", "Стать", "Дата народження",
                      "Посада", "Місто проживання", "Адреса проживання", "Телефон", "Email"]

        ws_all = wb.create_sheet(title="all")
        ws_all.append(["№", "Прізвище", "Ім’я", "По батькові", "Стать", "Дата народження", "Посада", "Місто проживання", "Адреса проживання", "Телефон", "Email"])

        ws_younger_18 = wb.create_sheet(title="younger_18")
        ws_younger_18.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

        ws_18_45 = wb.create_sheet(title="18-45")
        ws_18_45.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

        ws_45_70 = wb.create_sheet(title="45-70")
        ws_45_70.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

        ws_older_70 = wb.create_sheet(title="older_70")
        ws_older_70.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

        count_all = 0
        count_younger_18 = 0
        count_18_45 = 0
        count_45_70 = 0
        count_older_70 = 0

        for row in csv_reader:
            birthdate = datetime.strptime(row["Дата народження"], '%d.%m.%Y')
            age = calculate_age(birthdate)

            if age < 18:
                category = "younger_18"
                count_younger_18 += 1
                row_number = count_younger_18
            elif 18 <= age <= 45:
                category = "18-45"
                count_18_45 += 1
                row_number = count_18_45
            elif 45 < age <= 70:
                category = "45-70"
                count_45_70 += 1
                row_number = count_45_70
            else:
                category = "older_70"
                count_older_70 += 1
                row_number = count_older_70

            count_all += 1
            ws_all.append([count_all, row["Прізвище"], row["Ім’я"], row["По батькові"], row["Стать"], row["Дата народження"], row["Посада"], row["Місто проживання"], row["Адреса проживання"], row["Телефон"], row["Email"]])

            ws_category = wb[category]
            ws_category.append([row_number, row["Прізвище"], row["Ім’я"], row["По батькові"], row["Дата народження"], age])

        wb.save("employees.xlsx")
        print("Ok")
except FileNotFoundError:
    print("Повідомлення про відсутність, або проблеми при відкритті файлу CSV.")
except Exception as e:
    print("Повідомлення про неможливість створення XLSX файлу:", str(e))
